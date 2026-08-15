"""
数据存储管理模块
负责 JSON 数据的读写、Excel 导出、数据文件初始化、GitHub 持久化同步
"""

import base64
import hashlib
import json
import os
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
import requests
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .scoring import (
    get_criteria,
    get_total_score,
    get_groups,
    is_practical_group,
    normalize_group,
)

MODULE_VERSION = "2026-08-15-beijing-deductions-v2"

# 数据目录
DATA_DIR = Path(__file__).parent.parent / "data"

# 裁判信息文件
JUDGES_FILE = DATA_DIR / "judges.json"

# 评分记录文件（沿用原文件名，保留已有评分数据）
SCORE_FILES = {
    "答辩组": DATA_DIR / "scores_线上答辩.json",
    "实操组": DATA_DIR / "scores_甘肃线下实操.json",
    "北京线上实操组": DATA_DIR / "scores_北京线上实操.json",
}

# 删除标记保存在 score-data 分支；即使 main 或其他实例仍有旧缓存，
# 被删除的评分记录也不会在后续合并时重新出现。
SCORE_DELETIONS_FILE = DATA_DIR / "score_deletions.json"
SCORE_DELETIONS_REPO_PATH = "data/score_deletions.json"

# GitHub 仓库信息。评分数据写入独立分支，避免每次提交触发应用重新部署。
GITHUB_REPO = "xiaoxiao07/judge-scoring-app"
GITHUB_BRANCH = "main"
GITHUB_DATA_BRANCH = "score-data"
GITHUB_RAW_BASE = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/"
GITHUB_API_BASE = f"https://api.github.com/repos/{GITHUB_REPO}"
GITHUB_TIMEOUT = 15
GITHUB_SYNC_RETRIES = 24

_SCORE_DATA_LOCK = threading.RLock()
_INITIALIZED = False


class ScorePersistenceError(RuntimeError):
    """评分未能在远端完成持久化确认。"""


def _get_github_token() -> str:
    """从 Streamlit secrets 或环境变量获取 GitHub token。"""
    env_token = os.environ.get("GITHUB_TOKEN", "")
    if env_token:
        return env_token
    try:
        import streamlit as st

        return str(st.secrets.get("GITHUB_TOKEN", ""))
    except Exception:
        return ""


def _github_headers(token: str = "") -> dict:
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Cache-Control": "no-cache",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def _load_from_github(file_path: Path, repo_path: str) -> bool:
    """从 main 分支 raw 地址加载普通 JSON 列表文件。"""
    url = GITHUB_RAW_BASE + repo_path
    try:
        response = requests.get(
            url,
            headers={"Cache-Control": "no-cache"},
            params={"_": time.time_ns()},
            timeout=10,
        )
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list):
                _write_json(file_path, data)
                return True
    except Exception:
        pass
    return False


def _sync_to_github(file_path: Path, repo_path: str, commit_msg: str) -> bool:
    """同步普通 JSON 文件（目前仅用于裁判注册信息）。"""
    token = _get_github_token()
    if not token:
        return False

    url = f"{GITHUB_API_BASE}/contents/{repo_path}"
    try:
        content = file_path.read_text(encoding="utf-8")
        encoded = base64.b64encode(content.encode("utf-8")).decode("utf-8")
        headers = _github_headers(token)
        response = requests.get(
            url,
            headers=headers,
            params={"ref": GITHUB_BRANCH},
            timeout=GITHUB_TIMEOUT,
        )
        sha = response.json().get("sha") if response.status_code == 200 else None
        payload = {
            "message": commit_msg,
            "content": encoded,
            "branch": GITHUB_BRANCH,
        }
        if sha:
            payload["sha"] = sha
        response = requests.put(
            url,
            json=payload,
            headers=headers,
            timeout=GITHUB_TIMEOUT,
        )
        return response.status_code in (200, 201)
    except Exception:
        return False


def _sync_judges_to_github() -> bool:
    """同步裁判信息到 GitHub。"""
    return _sync_to_github(
        JUDGES_FILE,
        "data/judges.json",
        "Auto-sync: update judges info",
    )


def _read_json(file_path: Path) -> list:
    """读取 JSON 列表文件。"""
    if not file_path.exists():
        return []
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            return data if isinstance(data, list) else []
    except (json.JSONDecodeError, FileNotFoundError, OSError):
        return []


def _write_json(file_path: Path, data: list):
    """使用临时文件和原子替换写入 JSON，避免并发读取到半写文件。"""
    file_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = file_path.with_name(
        f".{file_path.name}.{os.getpid()}.{threading.get_ident()}.{uuid.uuid4().hex}.tmp"
    )
    try:
        with open(temp_path, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)
            file.flush()
            os.fsync(file.fileno())
        os.replace(temp_path, file_path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _legacy_record_digest(record: dict) -> str:
    payload = {key: value for key, value in record.items() if key != "record_id"}
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _ensure_record_ids(records: list) -> list:
    """为旧记录生成稳定 ID；相同旧记录通过出现序号保留，不误去重。"""
    occurrences = {}
    normalized = []
    for source_record in records:
        if not isinstance(source_record, dict):
            continue
        record = dict(source_record)
        if not record.get("record_id"):
            digest = _legacy_record_digest(record)
            occurrences[digest] = occurrences.get(digest, 0) + 1
            record["record_id"] = f"legacy-{digest[:24]}-{occurrences[digest]}"
        normalized.append(record)
    return normalized


def _merge_score_records(*record_sets: list) -> list:
    """按 record_id 合并多份记录；先出现的版本优先，任何记录都不会被整文件覆盖。"""
    merged = []
    seen_ids = set()
    for records in record_sets:
        for record in _ensure_record_ids(records or []):
            record_id = record["record_id"]
            if record_id in seen_ids:
                continue
            seen_ids.add(record_id)
            merged.append(record)
    return merged


def _normalize_score_deletions(deletions: list) -> list:
    """规范化删除标记，并按（评分组、提交ID）去重。"""
    normalized = []
    seen = set()
    for source_marker in deletions or []:
        if not isinstance(source_marker, dict):
            continue
        record_id = str(source_marker.get("record_id", "")).strip()
        group = normalize_group(str(source_marker.get("group", "")).strip())
        if not record_id or group not in SCORE_FILES:
            continue
        key = (group, record_id)
        if key in seen:
            continue
        marker = dict(source_marker)
        marker["group"] = group
        marker["record_id"] = record_id
        normalized.append(marker)
        seen.add(key)
    return normalized


def _score_deletion_keys(deletions: list) -> set:
    return {
        (marker["group"], marker["record_id"])
        for marker in _normalize_score_deletions(deletions)
    }


def _filter_deleted_score_records(group: str, records: list, deletions: list) -> list:
    """过滤已删除记录；返回值始终带稳定 record_id。"""
    normalized_group = normalize_group(group)
    deleted_ids = {
        record_id
        for marker_group, record_id in _score_deletion_keys(deletions)
        if marker_group == normalized_group
    }
    return [
        record
        for record in _ensure_record_ids(records or [])
        if record.get("record_id") not in deleted_ids
    ]


def _fetch_github_records(repo_path: str, branch: str, token: str = "") -> dict:
    """读取指定分支上的 JSON 记录文件，并返回内容及用于 CAS 的 SHA。"""
    url = f"{GITHUB_API_BASE}/contents/{repo_path}"
    try:
        response = requests.get(
            url,
            headers=_github_headers(token),
            params={"ref": branch},
            timeout=GITHUB_TIMEOUT,
        )
        if response.status_code == 404:
            return {"ok": True, "exists": False, "records": [], "sha": None, "error": ""}
        if response.status_code != 200:
            # 未配置令牌的只读场景可能触发 GitHub API 速率限制；
            # 后台读取可退回 raw 文件，但写入仍必须使用带 SHA 的认证 API。
            if not token:
                raw_url = (
                    f"https://raw.githubusercontent.com/{GITHUB_REPO}/"
                    f"{branch}/{repo_path}"
                )
                raw_response = requests.get(
                    raw_url,
                    headers={"Cache-Control": "no-cache"},
                    params={"_": time.time_ns()},
                    timeout=GITHUB_TIMEOUT,
                )
                if raw_response.status_code == 404:
                    return {
                        "ok": True,
                        "exists": False,
                        "records": [],
                        "sha": None,
                        "error": "",
                    }
                if raw_response.status_code == 200:
                    records = raw_response.json()
                    if isinstance(records, list):
                        return {
                            "ok": True,
                            "exists": True,
                            "records": records,
                            "sha": None,
                            "error": "",
                        }
            return {
                "ok": False,
                "exists": False,
                "records": [],
                "sha": None,
                "error": f"GitHub 读取失败（HTTP {response.status_code}）",
            }

        payload = response.json()
        if payload.get("encoding") == "base64" and payload.get("content"):
            raw = base64.b64decode(payload["content"]).decode("utf-8")
        elif payload.get("download_url"):
            download = requests.get(
                payload["download_url"],
                headers={"Cache-Control": "no-cache"},
                params={"_": time.time_ns()},
                timeout=GITHUB_TIMEOUT,
            )
            if download.status_code != 200:
                return {
                    "ok": False,
                    "exists": True,
                    "records": [],
                    "sha": payload.get("sha"),
                    "error": f"GitHub 文件下载失败（HTTP {download.status_code}）",
                }
            raw = download.text
        else:
            return {
                "ok": False,
                "exists": True,
                "records": [],
                "sha": payload.get("sha"),
                "error": "GitHub 返回的评分文件内容不可读取",
            }

        records = json.loads(raw)
        if not isinstance(records, list):
            raise ValueError("评分文件不是 JSON 列表")
        return {
            "ok": True,
            "exists": True,
            "records": records,
            "sha": payload.get("sha"),
            "error": "",
        }
    except Exception as exc:
        return {
            "ok": False,
            "exists": False,
            "records": [],
            "sha": None,
            "error": f"GitHub 读取异常：{exc}",
        }


def _put_github_records(
    repo_path: str,
    records: list,
    sha: Optional[str],
    token: str,
    commit_msg: str,
    branch: str = GITHUB_DATA_BRANCH,
) -> dict:
    """使用当前文件 SHA 更新指定分支；SHA 不匹配时由上层重试。"""
    url = f"{GITHUB_API_BASE}/contents/{repo_path}"
    content = json.dumps(records, ensure_ascii=False, indent=2)
    payload = {
        "message": commit_msg,
        "content": base64.b64encode(content.encode("utf-8")).decode("utf-8"),
        "branch": branch,
    }
    if sha:
        payload["sha"] = sha
    try:
        response = requests.put(
            url,
            json=payload,
            headers=_github_headers(token),
            timeout=GITHUB_TIMEOUT,
        )
        if response.status_code in (200, 201):
            return {"ok": True, "conflict": False, "status": response.status_code, "error": ""}
        message = ""
        try:
            message = response.json().get("message", "")
        except Exception:
            message = response.text[:200]
        return {
            "ok": False,
            "conflict": response.status_code in (409, 422),
            "status": response.status_code,
            "error": f"GitHub 写入失败（HTTP {response.status_code}）：{message}",
        }
    except Exception as exc:
        return {
            "ok": False,
            "conflict": True,
            "status": 0,
            "error": f"GitHub 写入异常：{exc}",
        }


def _ensure_data_branch(token: str) -> tuple[bool, str]:
    """确保独立评分数据分支存在。"""
    headers = _github_headers(token)
    data_ref_url = f"{GITHUB_API_BASE}/git/ref/heads/{GITHUB_DATA_BRANCH}"
    try:
        response = requests.get(data_ref_url, headers=headers, timeout=GITHUB_TIMEOUT)
        if response.status_code == 200:
            return True, ""
        if response.status_code != 404:
            return False, f"无法检查数据分支（HTTP {response.status_code}）"

        main_ref_url = f"{GITHUB_API_BASE}/git/ref/heads/{GITHUB_BRANCH}"
        response = requests.get(main_ref_url, headers=headers, timeout=GITHUB_TIMEOUT)
        if response.status_code != 200:
            return False, f"无法读取主分支（HTTP {response.status_code}）"
        main_sha = response.json()["object"]["sha"]

        response = requests.post(
            f"{GITHUB_API_BASE}/git/refs",
            headers=headers,
            json={"ref": f"refs/heads/{GITHUB_DATA_BRANCH}", "sha": main_sha},
            timeout=GITHUB_TIMEOUT,
        )
        if response.status_code == 201:
            return True, ""
        if response.status_code == 422:
            verify = requests.get(data_ref_url, headers=headers, timeout=GITHUB_TIMEOUT)
            if verify.status_code == 200:
                return True, ""
        return False, f"无法创建数据分支（HTTP {response.status_code}）"
    except Exception as exc:
        return False, f"检查数据分支异常：{exc}"


def _compact_score_records_on_branch(
    group: str,
    branch: str,
    deletions: list,
    token: str,
) -> str:
    """从指定分支的活动评分文件中移除已标记记录，返回空串表示成功。"""
    normalized_group = normalize_group(group)
    repo_path = f"data/{SCORE_FILES[normalized_group].name}"
    last_error = "评分文件清理未完成"
    for attempt in range(GITHUB_SYNC_RETRIES):
        snapshot = _fetch_github_records(repo_path, branch, token)
        if not snapshot["ok"]:
            last_error = snapshot["error"]
            _sleep_for_retry(attempt)
            continue
        current_records = _ensure_record_ids(snapshot["records"])
        remaining_records = _filter_deleted_score_records(
            normalized_group,
            current_records,
            deletions,
        )
        if len(remaining_records) == len(current_records):
            return ""
        result = _put_github_records(
            repo_path,
            remaining_records,
            snapshot["sha"],
            token,
            f"Admin: remove deleted {normalized_group} score records",
            branch=branch,
        )
        if result["ok"]:
            return ""
        last_error = result["error"]
        if not result["conflict"]:
            break
        _sleep_for_retry(attempt)
    return last_error


def delete_score_records(record_ids_by_group: dict) -> dict:
    """永久隐藏并清理选中的评分记录；删除标记使用 SHA 条件更新。"""
    target_keys = set()
    for group, record_ids in (record_ids_by_group or {}).items():
        normalized_group = normalize_group(group)
        if normalized_group not in SCORE_FILES:
            continue
        for record_id in record_ids or []:
            normalized_id = str(record_id).strip()
            if normalized_id:
                target_keys.add((normalized_group, normalized_id))
    if not target_keys:
        return {"deleted_count": 0, "cleanup_warnings": []}

    token = _get_github_token()
    if not token:
        raise ScorePersistenceError("未配置 GITHUB_TOKEN，无法删除云端评分记录")
    branch_ok, branch_error = _ensure_data_branch(token)
    if not branch_ok:
        raise ScorePersistenceError(branch_error)

    with _SCORE_DATA_LOCK:
        committed_deletions = None
        last_error = "删除标记写入未完成"
        for attempt in range(GITHUB_SYNC_RETRIES):
            snapshot = _fetch_github_records(
                SCORE_DELETIONS_REPO_PATH,
                GITHUB_DATA_BRANCH,
                token,
            )
            if not snapshot["ok"]:
                last_error = snapshot["error"]
                _sleep_for_retry(attempt)
                continue
            current_deletions = _normalize_score_deletions(snapshot["records"])
            current_keys = _score_deletion_keys(current_deletions)
            new_markers = [
                {
                    "group": group,
                    "record_id": record_id,
                    "deleted_at_utc": datetime.now(timezone.utc).isoformat(
                        timespec="milliseconds"
                    ),
                    "deleted_by": "admin",
                }
                for group, record_id in sorted(target_keys - current_keys)
            ]
            candidate = current_deletions + new_markers
            if not new_markers:
                committed_deletions = candidate
                break
            result = _put_github_records(
                SCORE_DELETIONS_REPO_PATH,
                candidate,
                snapshot["sha"],
                token,
                f"Admin: mark {len(new_markers)} score records deleted",
                branch=GITHUB_DATA_BRANCH,
            )
            if result["ok"]:
                committed_deletions = candidate
                break
            last_error = result["error"]
            if not result["conflict"]:
                break
            _sleep_for_retry(attempt)

        if committed_deletions is None:
            raise ScorePersistenceError(last_error)

        _write_json(SCORE_DELETIONS_FILE, committed_deletions)
        cleanup_warnings = []
        affected_groups = sorted({group for group, _ in target_keys})
        for group in affected_groups:
            for branch in (GITHUB_DATA_BRANCH, GITHUB_BRANCH):
                cleanup_error = _compact_score_records_on_branch(
                    group,
                    branch,
                    committed_deletions,
                    token,
                )
                if cleanup_error:
                    cleanup_warnings.append(f"{group}/{branch}：{cleanup_error}")

            file_path = SCORE_FILES[group]
            local_records = _filter_deleted_score_records(
                group,
                _read_json(file_path),
                committed_deletions,
            )
            _write_json(file_path, local_records)

        return {
            "deleted_count": len(target_keys),
            "cleanup_warnings": cleanup_warnings,
        }


def _sleep_for_retry(attempt: int):
    time.sleep(min(0.05 * (attempt + 1), 0.5))


def _persist_score_records_to_github(
    group: str,
    local_records: list,
    required_record_id: str,
) -> tuple[bool, list, str]:
    """
    将本地记录与 score-data、main 两个分支合并后以 SHA 条件更新。
    并发冲突会重新读取、合并并重试；删除标记优先于所有旧缓存。
    """
    token = _get_github_token()
    if not token:
        return False, local_records, "未配置 GITHUB_TOKEN，无法确认云端持久化"

    branch_ok, branch_error = _ensure_data_branch(token)
    if not branch_ok:
        return False, local_records, branch_error

    file_path = SCORE_FILES[group]
    repo_path = f"data/{file_path.name}"
    last_error = "GitHub 同步未完成"

    for attempt in range(GITHUB_SYNC_RETRIES):
        primary = _fetch_github_records(repo_path, GITHUB_DATA_BRANCH, token)
        legacy = _fetch_github_records(repo_path, GITHUB_BRANCH, token)
        deletion_snapshot = _fetch_github_records(
            SCORE_DELETIONS_REPO_PATH,
            GITHUB_DATA_BRANCH,
            token,
        )
        if not primary["ok"] or not legacy["ok"] or not deletion_snapshot["ok"]:
            last_error = (
                primary["error"]
                or legacy["error"]
                or deletion_snapshot["error"]
            )
            _sleep_for_retry(attempt)
            continue

        deletions = _normalize_score_deletions(deletion_snapshot["records"])
        _write_json(SCORE_DELETIONS_FILE, deletions)
        candidate = _filter_deleted_score_records(
            group,
            _merge_score_records(
                primary["records"],
                legacy["records"],
                local_records,
            ),
            deletions,
        )
        candidate_ids = {record["record_id"] for record in candidate}
        if required_record_id not in candidate_ids:
            return False, candidate, "该提交已被管理员删除，不能通过旧页面重新写入"

        primary_records = _filter_deleted_score_records(
            group,
            primary["records"],
            deletions,
        )
        primary_ids = {record["record_id"] for record in primary_records}
        if candidate_ids.issubset(primary_ids):
            return True, _merge_score_records(primary_records, candidate), ""

        result = _put_github_records(
            repo_path,
            candidate,
            primary["sha"],
            token,
            f"Score-sync: append {group} record {required_record_id[:12]}",
        )
        if result["ok"]:
            return True, candidate, ""

        last_error = result["error"]
        if not result["conflict"] and result["status"] in (400, 401, 403, 404):
            break
        _sleep_for_retry(attempt)

    final_snapshot = _fetch_github_records(repo_path, GITHUB_DATA_BRANCH, token)
    final_deletions = _fetch_github_records(
        SCORE_DELETIONS_REPO_PATH,
        GITHUB_DATA_BRANCH,
        token,
    )
    if final_snapshot["ok"] and final_deletions["ok"]:
        deletions = _normalize_score_deletions(final_deletions["records"])
        remote_records = _filter_deleted_score_records(
            group,
            final_snapshot["records"],
            deletions,
        )
        remote_ids = {record["record_id"] for record in remote_records}
        allowed_local = _filter_deleted_score_records(group, local_records, deletions)
        local_ids = {record["record_id"] for record in allowed_local}
        if required_record_id in remote_ids and local_ids.issubset(remote_ids):
            return True, remote_records, ""

    return False, local_records, last_error


def refresh_score_cache(group: str, require_remote: bool = False) -> list:
    """合并远端与本地缓存，并优先过滤管理员删除的记录。"""
    normalized_group = normalize_group(group)
    file_path = SCORE_FILES.get(normalized_group)
    if not file_path:
        return []

    token = _get_github_token()
    repo_path = f"data/{file_path.name}"
    with _SCORE_DATA_LOCK:
        local_records = _read_json(file_path)
        primary = _fetch_github_records(repo_path, GITHUB_DATA_BRANCH, token)
        legacy = _fetch_github_records(repo_path, GITHUB_BRANCH, token)
        deletion_snapshot = _fetch_github_records(
            SCORE_DELETIONS_REPO_PATH,
            GITHUB_DATA_BRANCH,
            token,
        )
        if require_remote and (
            not primary["ok"]
            or not legacy["ok"]
            or not deletion_snapshot["ok"]
        ):
            detail = (
                primary["error"]
                or legacy["error"]
                or deletion_snapshot["error"]
                or "远端历史记录读取失败"
            )
            raise ScorePersistenceError(detail)

        if deletion_snapshot["ok"]:
            deletions = _normalize_score_deletions(deletion_snapshot["records"])
            _write_json(SCORE_DELETIONS_FILE, deletions)
        else:
            deletions = _normalize_score_deletions(
                _read_json(SCORE_DELETIONS_FILE)
            )

        record_sets = []
        if primary["ok"]:
            record_sets.append(primary["records"])
        if legacy["ok"]:
            record_sets.append(legacy["records"])
        record_sets.append(local_records)
        merged = _filter_deleted_score_records(
            normalized_group,
            _merge_score_records(*record_sets),
            deletions,
        )
        _write_json(file_path, merged)
        return merged


def init_data_files():
    """初始化数据文件，并在每个进程首次启动时合并远端历史记录。"""
    global _INITIALIZED
    with _SCORE_DATA_LOCK:
        if _INITIALIZED:
            return
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if not JUDGES_FILE.exists():
            _write_json(JUDGES_FILE, [])
        if not SCORE_DELETIONS_FILE.exists():
            _write_json(SCORE_DELETIONS_FILE, [])
        for group in get_groups():
            if group in SCORE_FILES and not SCORE_FILES[group].exists():
                _write_json(SCORE_FILES[group], [])

        _load_from_github(JUDGES_FILE, "data/judges.json")
        for group in get_groups():
            if group in SCORE_FILES:
                refresh_score_cache(group, require_remote=False)
        _INITIALIZED = True


# ===================== 裁判管理 =====================


def register_judge(name: str, judge_id: str, group: str) -> dict:
    """注册或更新裁判；有令牌时使用 SHA 条件更新避免覆盖并发删除。"""
    normalized_group = normalize_group(group)
    judge_token = hashlib.sha256(f"{name}|{judge_id}".encode()).hexdigest()[:12]

    def upsert(judges: list) -> tuple[list, dict, bool]:
        normalized_judges = [item for item in judges if isinstance(item, dict)]
        for judge in normalized_judges:
            if str(judge.get("judge_id", "")) == str(judge_id):
                changed = (
                    judge.get("name") != name
                    or normalize_group(judge.get("group", "")) != normalized_group
                    or not judge.get("token")
                )
                judge["name"] = name
                judge["judge_id"] = judge_id
                judge["group"] = normalized_group
                if not judge.get("token"):
                    judge["token"] = judge_token
                return normalized_judges, judge, changed
        judge_info = {
            "name": name,
            "judge_id": judge_id,
            "group": normalized_group,
            "token": judge_token,
        }
        normalized_judges.append(judge_info)
        return normalized_judges, judge_info, True

    token = _get_github_token()
    if token:
        with _SCORE_DATA_LOCK:
            for attempt in range(GITHUB_SYNC_RETRIES):
                snapshot = _fetch_github_records(
                    "data/judges.json",
                    GITHUB_BRANCH,
                    token,
                )
                if not snapshot["ok"]:
                    _sleep_for_retry(attempt)
                    continue
                judges, judge_info, changed = upsert(snapshot["records"])
                if not changed:
                    _write_json(JUDGES_FILE, judges)
                    return judge_info
                result = _put_github_records(
                    "data/judges.json",
                    judges,
                    snapshot["sha"],
                    token,
                    f"Judge-sync: register or update {judge_id}",
                    branch=GITHUB_BRANCH,
                )
                if result["ok"]:
                    _write_json(JUDGES_FILE, judges)
                    return judge_info
                if not result["conflict"]:
                    break
                _sleep_for_retry(attempt)

    # 无令牌或远端暂不可用时保留原有本地登录能力，但不覆盖远端最新列表。
    judges, judge_info, _ = upsert(_read_json(JUDGES_FILE))
    _write_json(JUDGES_FILE, judges)
    return judge_info


def find_judge_by_token(token: str) -> Optional[dict]:
    """根据 token 查找裁判信息。"""
    for judge in _read_json(JUDGES_FILE):
        if judge.get("token") == token:
            return judge
    return None


def find_judge_by_id(judge_id: str) -> Optional[dict]:
    """根据编号查找裁判信息。"""
    for judge in _read_json(JUDGES_FILE):
        if judge["judge_id"] == judge_id:
            return judge
    return None


def get_all_judges(
    refresh_remote: bool = False,
    require_remote: bool = False,
) -> list:
    """获取已注册裁判；管理页可要求远端读取必须成功。"""
    if refresh_remote:
        snapshot = _fetch_github_records(
            "data/judges.json",
            GITHUB_BRANCH,
            _get_github_token(),
        )
        if snapshot["ok"]:
            judges = [item for item in snapshot["records"] if isinstance(item, dict)]
            _write_json(JUDGES_FILE, judges)
            return judges
        if require_remote:
            raise ScorePersistenceError(snapshot["error"] or "远端裁判列表读取失败")
    return _read_json(JUDGES_FILE)


def delete_judges(judge_keys: list) -> int:
    """按（裁判姓名、裁判编号）删除注册信息，不自动删除评分记录。"""
    targets = {
        (str(name).strip(), str(judge_id).strip())
        for name, judge_id in (judge_keys or [])
        if str(name).strip() or str(judge_id).strip()
    }
    if not targets:
        return 0
    token = _get_github_token()
    if not token:
        raise ScorePersistenceError("未配置 GITHUB_TOKEN，无法删除云端裁判信息")

    with _SCORE_DATA_LOCK:
        last_error = "裁判信息删除未完成"
        for attempt in range(GITHUB_SYNC_RETRIES):
            snapshot = _fetch_github_records(
                "data/judges.json",
                GITHUB_BRANCH,
                token,
            )
            if not snapshot["ok"]:
                last_error = snapshot["error"]
                _sleep_for_retry(attempt)
                continue
            current_judges = [
                item for item in snapshot["records"] if isinstance(item, dict)
            ]
            remaining = [
                judge
                for judge in current_judges
                if (
                    str(judge.get("name", "")).strip(),
                    str(judge.get("judge_id", "")).strip(),
                )
                not in targets
            ]
            removed_count = len(current_judges) - len(remaining)
            if removed_count == 0:
                local_remaining = [
                    judge
                    for judge in _read_json(JUDGES_FILE)
                    if (
                        str(judge.get("name", "")).strip(),
                        str(judge.get("judge_id", "")).strip(),
                    )
                    not in targets
                ]
                _write_json(JUDGES_FILE, local_remaining)
                return 0
            result = _put_github_records(
                "data/judges.json",
                remaining,
                snapshot["sha"],
                token,
                f"Admin: delete {removed_count} registered judges",
                branch=GITHUB_BRANCH,
            )
            if result["ok"]:
                _write_json(JUDGES_FILE, remaining)
                return removed_count
            last_error = result["error"]
            if not result["conflict"]:
                break
            _sleep_for_retry(attempt)

    raise ScorePersistenceError(last_error)


# ===================== 评分记录管理 =====================


def save_score(
    judge_info: dict,
    contestant_id: str,
    scores: dict,
    deductions: Optional[dict] = None,
    final_score: Optional[float] = None,
    veto_triggered: bool = False,
    veto_items: Optional[list] = None,
    score_zero_triggered: bool = False,
    score_zero_items: Optional[list] = None,
    score_overrides: Optional[list] = None,
    duration: Optional[str] = None,
    record_id: Optional[str] = None,
    contestant_group: Optional[str] = None,
) -> dict:
    """
    保存一条评分记录。仅当 GitHub 已确认包含该 record_id 时返回成功；
    同一个 record_id 可安全重试，不会产生重复记录。
    """
    group = normalize_group(judge_info["group"])
    file_path = SCORE_FILES.get(group)
    if not file_path:
        raise ValueError(f"未知的裁判组: {group}")

    record_id = record_id or uuid.uuid4().hex
    with _SCORE_DATA_LOCK:
        records = _ensure_record_ids(_read_json(file_path))
        existing = next(
            (record for record in records if record.get("record_id") == record_id),
            None,
        )
        if existing:
            record = existing
        else:
            score_total = sum(scores.values())
            record = {
                "record_id": record_id,
                "judge_name": judge_info["name"],
                "judge_id": judge_info["judge_id"],
                "judge_group": group,
                "contestant_id": contestant_id,
                "contestant_group": contestant_group or "",
                "scores": {
                    key: int(value)
                    if isinstance(value, float) and value.is_integer()
                    else value
                    for key, value in scores.items()
                },
                "total_score": final_score if final_score is not None else score_total,
                "raw_score": score_total,
                "total_max": get_total_score(group),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "saved_at_utc": datetime.now(timezone.utc).isoformat(timespec="milliseconds"),
            }
            if duration:
                record["duration"] = duration
            if deductions:
                record["deductions"] = deductions
                record["deduction_total"] = sum(
                    value
                    for value in deductions.values()
                    if isinstance(value, (int, float))
                )
            if veto_triggered:
                record["veto_triggered"] = True
                record["veto_items"] = veto_items or []
            if score_zero_triggered:
                record["score_zero_triggered"] = True
                record["score_zero_items"] = score_zero_items or []
            if score_overrides:
                record["score_overrides"] = score_overrides

        local_records = _merge_score_records(records, [record])
        success, confirmed_records, error = _persist_score_records_to_github(
            group,
            local_records,
            record_id,
        )
        if not success:
            raise ScorePersistenceError(
                f"云端未确认保存，系统没有显示成功：{error}。请保持当前页面并重试"
            )

        confirmed_records = _ensure_record_ids(confirmed_records)
        _write_json(file_path, confirmed_records)
        return next(
            confirmed
            for confirmed in confirmed_records
            if confirmed.get("record_id") == record_id
        )


def get_all_scores(
    group: str,
    refresh_remote: bool = False,
    require_remote: bool = False,
) -> list:
    """获取某组全部记录；后台导出可要求远端刷新必须成功。"""
    normalized_group = normalize_group(group)
    file_path = SCORE_FILES.get(normalized_group)
    if not file_path:
        return []
    if refresh_remote:
        return refresh_score_cache(normalized_group, require_remote=require_remote)
    with _SCORE_DATA_LOCK:
        deletions = _normalize_score_deletions(_read_json(SCORE_DELETIONS_FILE))
        records = _filter_deleted_score_records(
            normalized_group,
            _read_json(file_path),
            deletions,
        )
        _write_json(file_path, records)
        return records


# ===================== Excel 导出 =====================


def _style_header(ws, headers: list, row: int = 1):
    """设置表头样式。"""
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B7C9E2"),
        right=Side(style="thin", color="B7C9E2"),
        top=Side(style="thin", color="B7C9E2"),
        bottom=Side(style="thin", color="B7C9E2"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[row].height = 42


def _format_mapping(mapping: Optional[dict]) -> str:
    if not mapping:
        return ""
    return "；".join(f"{key}：{value}" for key, value in mapping.items())


def _format_items(items: Optional[list]) -> str:
    return "；".join(str(item) for item in (items or []))


def _apply_sheet_layout(ws, headers: list, width_cap: int = 30):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for col_idx, header_text in enumerate(headers, 1):
        text_width = sum(2 if ord(char) > 127 else 1 for char in str(header_text)) + 2
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
            max(text_width, 10),
            width_cap,
        )


def export_records_to_excel(
    group: str,
    records: list,
    file_label: str = "评分记录",
) -> Optional[Path]:
    """将给定评分记录完整导出为 Excel，不重新读取或修改远端数据。"""
    normalized_group = normalize_group(group)
    records = _ensure_record_ids(records or [])
    if not records:
        return None

    criteria = get_criteria(normalized_group)
    workbook = openpyxl.Workbook()
    score_sheet = workbook.active
    score_sheet.title = f"{normalized_group}评分记录"

    score_headers = [f"{key}({value['max']})" for key, value in criteria.items()]
    is_practical = is_practical_group(normalized_group)
    if is_practical:
        # 实操类组别固定列：A 选手编号、B 选手组别、C/D 空列、E 总分、F 时间。
        headers = [
            "选手编号",
            "选手组别",
            "",
            "",
            "总分",
            "时间",
            "裁判姓名",
            "裁判编号",
            "裁判组",
            "评分时间",
            "提交ID",
            "原始总分",
            "扣分合计",
            "满分",
        ] + score_headers
    else:
        headers = [
            "提交ID",
            "裁判姓名",
            "裁判编号",
            "裁判组",
            "选手编号/姓名",
            "选手组别",
        ] + score_headers + ["原始总分", "扣分合计", "最终总分", "满分", "评分时间"]
    _style_header(score_sheet, headers)

    data_font = Font(size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for row_idx, record in enumerate(records, 2):
        raw_score = record.get("raw_score", record.get("total_score", 0))
        deduction_total = record.get("deduction_total", 0)
        if is_practical:
            row_data = [
                record.get("contestant_id", ""),
                record.get("contestant_group", ""),
                "",
                "",
                record.get("total_score", 0),
                record.get("duration", ""),
                record.get("judge_name", ""),
                record.get("judge_id", ""),
                record.get("judge_group", normalized_group),
                record.get("timestamp") or record.get("saved_at_utc", ""),
                record.get("record_id", ""),
                raw_score,
                deduction_total if deduction_total else "",
                record.get("total_max", get_total_score(normalized_group)),
            ]
            row_data += [
                record.get("scores", {}).get(criterion_key, 0)
                for criterion_key in criteria
            ]
        else:
            row_data = [
                record.get("record_id", ""),
                record.get("judge_name", ""),
                record.get("judge_id", ""),
                record.get("judge_group", normalized_group),
                record.get("contestant_id", ""),
                record.get("contestant_group", ""),
            ]
            row_data += [
                record.get("scores", {}).get(criterion_key, 0)
                for criterion_key in criteria
            ]
            row_data += [
                raw_score,
                deduction_total if deduction_total else "",
                record.get("total_score", 0),
                record.get("total_max", get_total_score(normalized_group)),
                record.get("timestamp") or record.get("saved_at_utc", ""),
            ]
        for col_idx, value in enumerate(row_data, 1):
            cell = score_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

    _apply_sheet_layout(score_sheet, headers, width_cap=30)
    if is_practical:
        score_sheet.column_dimensions["A"].width = 20
        score_sheet.column_dimensions["B"].width = 18
        score_sheet.column_dimensions["C"].width = 4
        score_sheet.column_dimensions["D"].width = 4
        score_sheet.column_dimensions["E"].width = 12
        score_sheet.column_dimensions["F"].width = 14
        score_sheet.column_dimensions["G"].width = 16
        score_sheet.column_dimensions["H"].width = 16
        score_sheet.column_dimensions["I"].width = 12
        score_sheet.column_dimensions["J"].width = 20
        score_sheet.column_dimensions["K"].width = 34
    else:
        score_sheet.column_dimensions["A"].width = 34

    audit_sheet = workbook.create_sheet(f"{normalized_group}提交审计")
    record_jsons = [
        json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for record in records
    ]
    json_chunk_size = 30000
    max_json_chunks = max(
        1,
        max((len(value) + json_chunk_size - 1) // json_chunk_size for value in record_jsons),
    )
    audit_headers = [
        "提交ID",
        "裁判姓名",
        "裁判编号",
        "裁判组",
        "选手编号/姓名",
        "选手组别",
        "用时",
        "扣分明细",
        "自动归零说明",
        "总分归零",
        "总分归零项",
        "否决触发",
        "否决项",
        "评分时间",
    ] + [f"完整记录JSON-{index + 1}" for index in range(max_json_chunks)]
    _style_header(audit_sheet, audit_headers)
    json_start_col = len(audit_headers) - max_json_chunks + 1

    for row_idx, (record, record_json) in enumerate(zip(records, record_jsons), 2):
        chunks = [
            record_json[start : start + json_chunk_size]
            for start in range(0, len(record_json), json_chunk_size)
        ] or [""]
        chunks += [""] * (max_json_chunks - len(chunks))
        row_data = [
            record.get("record_id", ""),
            record.get("judge_name", ""),
            record.get("judge_id", ""),
            record.get("judge_group", normalized_group),
            record.get("contestant_id", ""),
            record.get("contestant_group", ""),
            record.get("duration", ""),
            _format_mapping(record.get("deductions")),
            _format_items(record.get("score_overrides")),
            "是" if record.get("score_zero_triggered") else "否",
            _format_items(record.get("score_zero_items")),
            "是" if record.get("veto_triggered") else "否",
            _format_items(record.get("veto_items")),
            record.get("timestamp") or record.get("saved_at_utc", ""),
        ] + chunks
        for col_idx, value in enumerate(row_data, 1):
            cell = audit_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = Alignment(
                horizontal="left" if col_idx >= 8 else "center",
                vertical="top",
                wrap_text=col_idx >= 8,
            )
            cell.border = thin_border
        audit_sheet.row_dimensions[row_idx].height = 42

    _apply_sheet_layout(audit_sheet, audit_headers, width_cap=38)
    audit_sheet.column_dimensions["A"].width = 34
    for col_idx in range(8, json_start_col):
        audit_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 38
    # 原始 JSON 用于完整取证，默认隐藏以保持审计表可读；需要时可在 Excel 中取消隐藏。
    for col_idx in range(json_start_col, len(audit_headers) + 1):
        dimension = audit_sheet.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ]
        dimension.width = 20
        dimension.hidden = True

    safe_label = "".join(
        "_" if char in '\\/:*?"<>|' else char
        for char in str(file_label)
    ).strip(" ._") or "评分记录"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    file_path = DATA_DIR / f"{normalized_group}_{safe_label}_{timestamp}.xlsx"
    workbook.save(file_path)
    return file_path


def export_to_excel(group: str) -> Optional[Path]:
    """刷新并导出某组全部历史评分；远端读取失败时拒绝生成不完整文件。"""
    normalized_group = normalize_group(group)
    records = get_all_scores(
        normalized_group,
        refresh_remote=True,
        require_remote=True,
    )
    return export_records_to_excel(normalized_group, records)


def export_all_to_excel() -> dict:
    """导出所有组的完整历史记录，返回 {组名: 文件路径}。"""
    results = {}
    for group in get_groups():
        file_path = export_to_excel(group)
        if file_path:
            results[group] = file_path
    return results
